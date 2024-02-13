
# import modules
import openpyxl
import csv
import os

contact_list = {}

def add_contact(name, number, email):
    if name not in contact_list:
        contact_list[name] = {'number': number, 'email': email}
        print("Contact added successfully!")
    else:
        print("Contact already exists!")


def search_contact(name):
    if name in contact_list:
        print(
            f'Contact found - Name: {name}, Number: {contact_list[name]["number"]}, Email: {contact_list[name]["email"]}')
    else:
        print('Contact not found')


def view_contacts():
    if contact_list:
        sort_contacts = sorted(contact_list.items(), key=lambda x: x[0])
        for name, contact_info in enumerate(sort_contacts, start=1):

            print(f'{name}. Name: {name}, Number: {number}, Email: {email}')
    else:
        print('No contacts available.')


def delete_contact(name):
    if name in contact_list:
        del contact_list[name]
        print(f"Contact:{name} deleted successfully!")
    else:
        print("Contact not found.")

def export_contacts(filename):

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet['A1'] = 'Name'
    sheet['B1'] = 'number'
    sheet['C1'] = 'email'
    for row , (name, contact_info) in enumerate(contact_list.items(),start=2):
        sheet[f'A{row}'] = name
        sheet[f'B{row}'] = contact_info['number']
        sheet[f'C{row}'] = contact_info['email']
        workbook.save(filename)
# automatically open the file
        open_file_command = f' start excel "{filename}"'
        os.system(open_file_command)
def import_file(filename):
    if filename.endswith('.xlsx'):
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            name, number, email = row
            add_contact(name,number,email)
            print("Contacts imported successfully!")
    elif filename.endswith('.csv'):
        with open(filename, 'r') as file:
            reader = csv.reader(file)
            next(reader)  # Skip the header row
            for row in reader:
                name, number, email = row
                add_contact(name, number, email)
    else:
        print('unsupported file format!')

while True:
    print('Welcome to your address book!')
    print('1. Add new contacts')
    print('2. Search contacts')
    print('3. Delete contacts')
    print('4. Show current contacts')
    print('5. Export contacts')
    print('6. Exit')
    print('7. Import contacts from Excel file format')


    choice = input('Choose your option below: ')

    if choice == '1':
        while True:
            name = input('Enter the name: ')
            number = input('Enter the number: ')
            email = input('Enter the email address: ')
            add_contact(name, number, email)
            new = input('would you like to add other names? (y/n)')
            if new != 'y':
                break
    elif choice == '2':
        name = input('Enter the name you are searching for: ')
        search_contact(name)
    elif choice == '3':
        name = input('Enter the name to delete: ')
        delete_contact(name)
    elif choice == '4':
        view_contacts()
    elif choice == '5':
        filename = input('enter the name you wish to save the file you will export: ')
        export_contacts(filename)
    elif choice == '6':
        print('Exiting the app.')
    elif choice == '7':
        filename = input('enter the name of the file you wish to import')
        import_file(filename)
        print('Exiting the app.')
        break
    else:
        print('Invalid choice.')




